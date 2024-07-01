import re
from datetime import datetime, date
import pandas as pd
from fpdf import FPDF
from openpyxl import load_workbook



class Person:
    def __init__(self, name, dob, phone, email, age):
        self.name = name
        self.dob = dob
        self.phone = phone
        self.email = email
        self.age = age
    
        
# Function to validate email
def validate_email(email):
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(pattern, email) is not None

# Function to validate phone number (10 digits)
def validate_phone(phone):
    pattern = r'^\d{10}$'
    return re.match(pattern, phone) is not None

# Function to validate date of birth (in the format YYYY-MM-DD)
def validate_dob(dob):
    try:
        datetime.strptime(dob, '%Y-%m-%d')
        return True
    except ValueError:
        return False

# Function to calculate age
def calculate_age(dob):
    birth_date = datetime.strptime(dob, '%Y-%m-%d').date()
    today = date.today()
    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
    return age

# Function to write to a text file
def write_to_text_file(data):
    with open("employee_data.txt", "w") as file:
        file.write(data)
    print("Data written to employee_data.txt")

# Function to write to a PDF file
def write_to_pdf_file(data):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.multi_cell(0, 10, data)
    pdf.output("employee_data.pdf")
    print("Data written to employee_data.pdf")

# Function to write to an Excel file
def write_to_excel_file(data):
    try:
        # Load existing workbook in append mode
        wb = load_workbook("employee_data.xlsx")
        sheet = wb.active

        # Determine the next available row
        next_row = sheet.max_row + 1

        # Write data to Excel sheet
        sheet.cell(row=next_row, column=1).value = data["Name"]
        sheet.cell(row=next_row, column=2).value = data["Date of Birth"]
        sheet.cell(row=next_row, column=3).value = data["Phone"]
        sheet.cell(row=next_row, column=4).value = data["E-Mail"]
        sheet.cell(row=next_row, column=5).value = data["Age"]

        # Save workbook
        wb.save("employee_data.xlsx")
        print("Data appended to employee_data.xlsx")

    except FileNotFoundError:
        # If file doesn't exist, create a new one and write data
        df = pd.DataFrame([data])
        df.to_excel("employee_data.xlsx", index=False)
        print("Data written to employee_data.xlsx")

# Function to get employee details from user
def get_employee_details():
    name = input("Enter Name: ")
    dob = input("Enter Date of Birth (YYYY-MM-DD): ")
    while not validate_dob(dob):
        print("Invalid date format. Please enter the date in YYYY-MM-DD format.")
        dob = input("Enter Date of Birth (YYYY-MM-DD): ")
    phone = input("Enter Phone (10 digits): ")
    while not validate_phone(phone):
        print("Invalid phone number. Please enter a 10-digit phone number.")
        phone = input("Enter Phone (10 digits): ")
    email = input("Enter E-Mail: ")
    while not validate_email(email):
        print("Invalid email format. Please enter a valid email address.")
        email = input("Enter E-Mail: ")
    age = calculate_age(dob)
    employee = {
        "Name": name,
        "Date of Birth": dob,
        "Phone": phone,
        "E-Mail": email,
        "Age": age
    }
    return employee

# Function to bulk read employee details from an Excel file
def bulk_read_from_excel(file_path):
    df = pd.read_excel(file_path)
    employees = df.to_dict(orient='records')
    for employee in employees:
        employee["Age"] = calculate_age(employee["Date of Birth"])
    return employees

# Main function to run the application
def main():
    print("Welcome to the Employee Management System")
    print("1. Enter Employee Details")
    print("2. Bulk Read from Excel")
    choice = input("Enter your choice: ").strip()

    if choice == "1":
        employee = get_employee_details()
        print("Choose the format to save the data:")
        print("1. Text")
        print("2. PDF")
        print("3. Excel")
        save_choice = input("Enter your choice: ").strip()
        
        data_str = (f"Name: {employee['Name']}\n"
                    f"Date of Birth: {employee['Date of Birth']}\n"
                    f"Phone: {employee['Phone']}\n"
                    f"E-Mail: {employee['E-Mail']}\n"
                    f"Age: {employee['Age']}\n")

        if save_choice == "1":
            write_to_text_file(data_str)
        elif save_choice == "2":
            write_to_pdf_file(data_str)
        elif save_choice == "3":
            write_to_excel_file(employee)
        else:
            print("Invalid choice. Exiting.")
    elif choice == "2":
        file_path = input("Enter the path of the Excel file: ").strip()
        employees = bulk_read_from_excel(file_path)
        for employee in employees:
            print(employee)
        # Save bulk data to an Excel file
        df = pd.DataFrame(employees)
        df.to_excel("bulk_employee_data.xlsx", index=False)
        print("Bulk data written to bulk_employee_data.xlsx")
    else:
        print("Invalid choice. Exiting.")

if __name__ == "__main__":
    main()
